# Nicholas Sylvester
# CS3080 Python Programming
# Task 1

"""Task 1 - Simple Word Game.

This program will scrape a website,
https://www.ef.edu/english-resources/english-vocabulary/top-1000-words/,
for the top 1000 English words.
The program will take these words and put them into an excel sheet.
This excel sheet will then be used as input to createa hangman-like
word game where the user has a limited number of guesses to guess
each letter in the random word out of the excel sheet.

"""

import requests
import bs4
import openpyxl
import random
import textwrap

FILE = "Task1.xlsx"
INCORRECT_GUESSES = 3


class Game():
    """Represents the game that the user plays."""
    
    def __init__(self):
        """Each game has data about the word and the guesses."""
        self.word = self.select_word(FILE)
        self.word_length = len(self.word)
        self.letters_guessed = []
        self.incorrect_guesses = 0
        self.guess_index = 1
        self.guessed_word = False

    def select_word(self, file):
        """Selecet a random word from word_bank."""
        wb = openpyxl.load_workbook(file)
        sheet = wb['Sheet']

        word = [cell.value for cell in sheet['A'] if cell.value is not None]

        selected_word = random.choice(word)
        return selected_word

    def display_word(self):
        """Display the partially reealed word."""
        for char in self.word:
            if char is not None:
                if char in self.letters_guessed:
                    print(f"{char}", end=' ')
                else:
                    print("_", end=' ')
        print()

    def guess_letter(self):
        """Get user input for a letter guess."""
        valid = False

        # Keep asking for a valid guess, doing some error checking first
        # Then determine if the letter has been guessed already
        # Then determine if the guess is in the word or not
        while not valid:
            guess = input(f"Guess {self.guess_index} (Incorrect Guesses {self.incorrect_guesses}): ").lower()
            if len(guess) == self.word_length:
                if guess == self.word:
                    print("You guessed the word!")
                    self.guessed_word is True
                    break
                else:
                    print(f"The word is not {guess}")
                    self.guess_index += 1
                    self.incorrect_guesses += 1
                    break
            if len(guess) > 1 or len(guess) < 1:
                print("ERROR: Please input only one letter or length of word")
                continue
            if guess.isdigit():
                print("ERROR: Guess must be a letter")
                continue
            if guess in self.letters_guessed:
                print(f"You have already guessed {guess}!")
                continue
            if guess not in self.word:
                print(f"There is no {guess} in the word")
                self.incorrect_guesses += 1
                self.guess_index += 1
                self.letters_guessed.append(guess)
                valid = True
            else:
                print(f"{guess} is in the word!")
                self.letters_guessed.append(guess)
                self.guess_index += 1
                valid = True


def get_words():
    """Get the words off of the internet."""
    try:
        response = requests.get('https://www.ef.edu/english-resources/english-vocabulary/top-1000-words/')
        response.raise_for_status()
    except Exception as e:
        # If requests can not connect to the link, display error and quit program
        # Functionality could be made to check if the user already has the sheet
        print(f"ERROR: {e}\nCheck Internet Connection")
        quit()

    # Create an html parser
    soup = bs4.BeautifulSoup(response.text, 'html.parser')

    # Find the div that contains the words
    div = soup.find('div', {'class': 'field-item even'})

    # Grab the paragraphs in the div
    paragraphs = div.find_all('p')

    # The words are stored in the 2nd <p> in <div>
    words = paragraphs[1].text
    # Put the words into a list
    words = words.split()

    return words


def create_sheet(words):
    """Create the word bank spreadsheet."""
    wb = openpyxl.Workbook()
    sheet = wb.active

    # Add words to the 'A' Column of the sheet
    for row, word in enumerate(words, start=1):
        sheet.cell(row, 1, word)

    wb.save('Task1.xlsx')


def main_game(game):
    """Play Hangman Game."""
    print("Welcome to Nick and Devon's Hangman Game!")
    print(textwrap.fill(f"You will have {INCORRECT_GUESSES} guesses to reveal the hidden word selected from the top 1000 English words.", 40))

    # Jump out of the loop if the player reached the max number of
    # incorrect guesses or if the word is guessed
    while game.incorrect_guesses < INCORRECT_GUESSES and not game.guessed_word:
        game.display_word()
        game.guess_letter()
        if all(char in game.letters_guessed for char in game.word):
            game.guessed_word = True

    # Basic end game to display if the user won or lost
    if not game.guessed_word:
        print("\nGame Over!")
        print(f"The word was {game.word}")
    else:
        print(f"\nYou correctly guessed the word, {game.word} in {game.guess_index} guesses!")


def main():
    """Run Main Function."""
    word_bank = get_words()
    create_sheet(word_bank)

    game = Game()

    main_game(game)


if __name__ == "__main__":
    main()
