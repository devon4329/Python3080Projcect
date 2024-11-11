"""
CS3080 Python - Project 1 - Task 2
This program will scrape three websites, https://openaccess.thecvf.com/CVPR2022?day=all,
https://openaccess.thecvf.com/CVPR2023?day=all, and https://openaccess.thecvf.com/CVPR2024?day=all,
and obtain the top three contributors (authors or researchers) in a conference for the last three years.
(2022, 2023, and 2024). The results will be saved into an Excel spreadsheet that shows how many times each
author contributed to a paper separated by the year.
"""
from venv import create

import openpyxl
from numpy.core.defchararray import center
from openpyxl.styles import Alignment, Font
import bs4
import requests

LINK_2022 = r'https://openaccess.thecvf.com/CVPR2022?day=all'
LINK_2023 = r'https://openaccess.thecvf.com/CVPR2023?day=all'
LINK_2024 = r'https://openaccess.thecvf.com/CVPR2024?day=all'


def find_contributors(link):
    try:
        response = requests.get(link)
        response.raise_for_status()

    except Exception as e:
        print(f"Error {e}\nCheck Internet Connection")
        quit()

    soup = bs4.BeautifulSoup(response.text, 'html.parser')

    div = soup.find('div', {'id' : 'content'})

    authors = list(div.select('input[type=hidden]'))

    author_counts = {}

    for element in authors:
        value = element.get("value")

        if value in author_counts:
            author_counts[value] += 1
        else:
            author_counts[value] = 1

    return author_counts

# Counts Totals of all lists
def count_totals(dict1, dict2, dict3):
    the_total = {}

    def add_totals(key, values):

        if key in the_total:
            the_total[key][0].append(values)
            the_total[key][1] += values
        else:
            the_total[key] = [[values], values]

    for dict_in in [dict1, dict2, dict3]:
        for key, value in dict_in.items():
            add_totals(key, value)

    sorted_list = sorted_list = sorted([(key, values[0], values[1])
                                        for key, values in the_total.items()], key=lambda x: x[2], reverse=True)

    return sorted_list



def create_sheet(aList):
    top_con = openpyxl.Workbook()
    print(top_con.sheetnames)
    sheet = top_con.active
    sheet.title = "CVF Top 3 Contributors"
    print(top_con.sheetnames)

    # Add years and total to sheet
    start_year = 2022
    for i in range(4):
        if i < 3:
            cell2022 = sheet.cell(row=i+2,column=1, value = start_year + i )
            cell2022.font = Font(bold=True)
            cell2022.alignment = Alignment(horizontal="center", vertical="center")
        else:
            cell2022 = sheet.cell(row=i+2, column=1, value = "Total")
            cell2022.font = Font(bold=True)
            cell2022.alignment = Alignment(horizontal="center")


    # Add names
    for i in range(3):
        cell = sheet.cell(row= 1, column = i+2, value = aList[i][0])
        cell.alignment = Alignment(horizontal="center", vertical="center")

    #Add year 2022 values to sheet
    index = 2
    for key in aList[:3]:
        i = 0
        for value in (key[1]):
            cell2 = sheet.cell(row=i+2, column=index, value=value)
            cell2.alignment = Alignment(horizontal="center", vertical="center")
            i += 1

        cell2 = sheet.cell(row=5, column = index, value=key[2])
        cell2.alignment = Alignment(horizontal="center", vertical="center")
        index += 1

    top_con.save('Task2.xlsx')


def main():
    contributors_2022 = find_contributors(LINK_2022)
    contributors_2023 = find_contributors(LINK_2023)
    contributors_2024 = find_contributors(LINK_2024)

    totals = count_totals(contributors_2022, contributors_2023, contributors_2024)

    create_sheet(totals)



if __name__ == "__main__":
    main()

